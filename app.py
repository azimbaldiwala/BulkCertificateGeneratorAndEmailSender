from PIL import Image, ImageDraw, ImageFont
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

file = openpyxl.load_workbook('Certi-data.xlsx')
sheet = file['Sheet1']

total_rows = sheet.max_row
#  getting data from excel sheet

names = []
countries = []
emails = []
dates = []

for i in range(2, total_rows + 1):
    names.append(sheet.cell(i, 1).value)
    countries.append(sheet.cell(i, 2).value)
    emails.append(sheet.cell(i, 4).value)
    t_date = str(sheet.cell(i, 3).value)
    t_date = t_date[:10]
    dates.append(t_date)

print("getting login credentials...")
email_file = open('emails.txt', 'r')
login_cred = email_file.read().split(sep=',')

email_id = login_cred[0]
email_id_password = login_cred[1]

email_user = email_id
email_password = email_id_password
server = smtplib.SMTP('smtp.gmail.com',587)
server.starttls()
print('logging in to the mail...')
server.login(email_user, email_password)

print('ready to generate certificates and sending mails...')
for (i, j, k, l) in zip(names, countries, dates, emails):
    print(f"Generating Certificate for {i}")
    im = Image.open('F_Inc_Certi.jpg')
    d = ImageDraw.Draw(im)
    name = (2020, 1200)
    country = (2020, 1485)
    date = (2020, 1765)
    text_color = (0, 0, 0)
    font = ImageFont.truetype("arial.ttf", 50)
    d.text(name, i, fill=text_color, font=font)
    d.text(country, j, fill=text_color, font=font)
    d.text(date, k, fill=text_color, font=font)
    im.save("certificate_" + i + ".pdf")
    print("Certificate generated successfully")
    print(f"Sending mail to Name: {i}, E-mail: {l}")
    email_send = l

    subject = 'Your certificate is ready.'

    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = email_send
    msg['Subject'] = subject

    body = 'Collect your certificate'
    msg.attach(MIMEText(body, 'plain'))
    filename = "certificate_" + i + ".pdf"
    attachment = open(filename, 'rb')

    part = MIMEBase('application','octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= "+filename)

    msg.attach(part)
    text = msg.as_string()



    server.sendmail(email_user, email_send, text)
    print(f"Mail sent to Name = {i}, Email = {l}")

print('closing server connection...')
server.quit()
print('job done successfully!')
